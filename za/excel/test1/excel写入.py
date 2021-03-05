#%%

from openpyxl import load_workbook

wb2 = load_workbook("1.xlsx")

print(wb2.sheetnames)

#%%
ws = wb2.active

ws['A4'] = 10
c=ws['A4'].value
print(c)

d=ws.cell(1,2,1000)
print(d.value)
wb2.save("2.xlsx")
# %%
from openpyxl import load_workbook

wb2 = load_workbook("1.xlsx")
ws = wb2.active

with open('1.txt') as f:
    line = f.read()
    data = line.split(" ")

k=0
for j in range(8):
    for i in range(16):
        d=ws.cell(6+j,5+i,data[k])
        k+=1
wb2.save("3.xlsx")
# %%
with open('3.txt') as f:
    line = f.read()
    data = line.split(" ")

with open('save.txt','w') as f:
    for d in data:
        f.write(",0x"+d)