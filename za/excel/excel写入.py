#%%

from openpyxl import load_workbook

wb2 = load_workbook("1.xlsx")

print(wb2.sheetnames)

#%%
ws = wb2.active

ws['A4'] = 10
c=ws['A4'].value
print(c)

d=ws.cell(1,2,1000)
print(d.value)
wb2.save("2.xlsx")
# %%
from openpyxl import load_workbook

wb2 = load_workbook("1.xlsx")
ws = wb2.active

with open('3.txt') as f:
    line = f.read()
    data = line.split(" ")

k=0
for j in range(8):
    for i in range(16):
        d=ws.cell(6+j,5+i,data[k])
        k+=1
wb2.save("4.xlsx")
# %%
with open('mac.txt') as f:
    line = f.read()
    data = line.split(" ")

with open('save.txt','w') as f:
    i=0
    for d in data:
        i+=1
        if i % 16 == 0:
            f.write("\n")
        if d != "":
            f.write("0x"+d+",")
        else:
            print("NG")
        # f.write(",0x"+d)
# %%
len(data)
# %%
