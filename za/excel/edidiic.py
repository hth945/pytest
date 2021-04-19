#%%
from openpyxl import load_workbook

wb2 = load_workbook("edidIIC.xlsx")

print(wb2.sheetnames)
ws = wb2[wb2.sheetnames[0]]
# %%
row = 1
liftEdid1=""
liftEdid2=""
rightEdid1=""
rightEdid2=""
while True:
    d=ws.cell(row,1)
    if d.value==None:
        break
    # if int(d.value,16)  == 0x0227: # 左
    #     liftEdid1+= ' ' + hex(int(ws.cell(row,2).value,16))[2:]  
    #     liftEdid2+= ',' + hex(int(ws.cell(row,2).value,16))
    # elif int(d.value,16)  == 0x0C27: # 左
    #     rightEdid1+= ' ' + hex(int(ws.cell(row,2).value,16))[2:]
    #     rightEdid2+= ',' + hex(int(ws.cell(row,2).value,16))

    if int(d.value,16)  == 0x0227: # 左
        liftEdid1+= ' {:02x}'.format(int(ws.cell(row,2).value,16))
        liftEdid2+= '0x{:02x},'.format(int(ws.cell(row,2).value,16))
    elif int(d.value,16)  == 0x0C27: # 左
        rightEdid1+= ' {:02x}'.format(int(ws.cell(row,2).value,16))
        rightEdid2+= '0x{:02x},'.format(int(ws.cell(row,2).value,16))


    row += 1
print(liftEdid1)
print(liftEdid2)
print(rightEdid1)
print(rightEdid2)
# %%

if liftEdid1 == rightEdid1:
    print(123)
# %%


DP_EDIDRegBuf= [0x0220,0x0226,0x022D,0x0C20,0x0C26,0x0C2D,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0227,0x0C27,0x0220,0x0C20,0x0931]

DP_EDIDDatBuf=[0x80,0x00,0x00,0x80,0x00,0x00,0x00,0x00,0xFF,0xFF,0xFF,0xFF,0xFF,0xFF,0xFF,0xFF,0xFF,0xFF,0xFF,0xFF,0x00,0x00,0x06,0x06,0x10,0x10,0x2D,0x2D,0xBF,0xBF,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x16,0x16,0x1C,0x1C,0x01,0x01,0x04,0x04,0xC5,0xC5,0x23,0x23,0x27,0x27,0x78,0x78,0x22,0x22,0x0F,0x0F,0x91,0x91,0xAE,0xAE,0x52,0x52,0x43,0x43,0xB0,0xB0,0x26,0x26,0x0F,0x0F,0x50,0x50,0x54,0x54,0x00,0x00,0x00,0x00,0x00,0x00,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x01,0x00,0x00,0x00,0x00,0x00,0x00,0x10,0x10,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x10,0x10,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0xFC,0xFC,0x00,0x00,0x58,0x58,0x31,0x31,0x30,0x30,0x36,0x36,0x33,0x33,0x20,0x20,0x45,0x45,0x78,0x78,0x70,0x70,0x2E,0x2E,0x0A,0x0A,0x20,0x20,0x20,0x20,0x00,0x00,0x00,0x00,0x00,0x00,0x10,0x10,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x01,0x01,0x94,0x94,0x70,0x70,0x13,0x13,0x79,0x79,0x03,0x03,0x00,0x00,0x12,0x12,0x00,0x00,0x16,0x16,0x82,0x82,0x10,0x10,0x00,0x10,0x00,0x00,0xFF,0xFF,0x09,0x09,0x3F,0x3F,0x0B,0x0B,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x41,0x41,0x50,0x50,0x50,0x50,0x2D,0x2D,0xBF,0xBF,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x03,0x03,0x00,0x00,0x14,0x14,0xCF,0xCF,0xB6,0xB6,0x00,0x00,0x88,0x88,0xFF,0xFF,0x09,0x09,0x27,0x27,0x00,0x00,0x03,0x03,0x80,0x80,0x0F,0x0F,0x00,0x00,0x3F,0x3F,0x0B,0x0B,0x77,0x77,0x00,0x00,0x69,0x69,0x00,0x00,0x07,0x07,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x82,0x72,0x90,0x90,0x00,0x00,0x03]

print(len(DP_EDIDRegBuf))
print(len(DP_EDIDDatBuf))
for i,data in enumerate(DP_EDIDRegBuf):
    d=int(ws.cell(i+1,1).value,16)
    if d != data:
        print(i,'{:04x}'.format(data),'{:04x}'.format(d))

for i,data in enumerate(DP_EDIDDatBuf):
    d=int(ws.cell(i+1,2).value,16)
    if d != data:
        print(i,'{:04x}'.format(data),'{:04x}'.format(d))

        

# %%
print(DP_EDIDRegBuf[1])
# %%
edidH=' 00 ff ff ff ff ff ff 00 06 10 2d bf 00 00 00 00 16 1c 01 04 c5 23 27 78 22 0f 91 ae 52 43 b0 26 0f 50 54 00 00 00 01 01 01 01 01 01 01 01 01 01 01 01 01 01 01 01 00 00 00 10 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 10 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 fc 00 58 31 30 36 33 20 45 78 70 2e 0a 20 20 00 00 00 10 00 00 00 00 00 00 00 00 00 00 00 00 00 00 01 94 70 13 79 03 00 12 00 16 82 10 00 00 ff 09 3f 0b 00 00 00 00 00 41 50 50 2d bf 00 00 00 00 03 00 14 cf b6 00 88 ff 09 27 00 03 80 0f 00 3f 0b 77 00 69 00 07 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 82 90'

if liftEdid1 == edidH:
    print(123)
else:
    print('ng')


#%%
with open('mac.txt') as f:
    line = f.read()
    data = line.split(" ")
dataStr = ''
dd = []
with open('save.txt','w') as f:
    i=0
    for d in data:
        i+=1
        if i % 16 == 0:
            f.write("\n")
        if d != "":
            f.write("0x"+d+",")
            dd.append(int(d,16))
            dataStr += "0x"+d+","
        else:
            print("NG")

# %%

with open('EDID_dp835_5120x2880_left_20201111_V01.txt') as f:
    lines = f.readlines()
fData=[]
for i in lines:
    fData.append(int(i,16))


data = liftEdid1.split(" ")
data = filter(lambda x : x, data)
dData=[]
for i in data:
    dData.append(int(i,16))

print(len(fData))
print(len(dData))
for t,d in enumerate(dData):
    if d != fData[t]:
        print('NG')
print('end')
# %%
with open('EDID_dp835_5120x2880_right_20201111_V01.txt') as f:
    lines = f.readlines()
fData=[]
for i in lines:
    fData.append(int(i,16))


data = rightEdid1.split(" ")
data = filter(lambda x : x, data)
dData=[]
for i in data:
    dData.append(int(i,16))

print(len(fData))
print(len(dData))
for t,d in enumerate(dData):
    if d != fData[t]:
        print('NG')
print('end')